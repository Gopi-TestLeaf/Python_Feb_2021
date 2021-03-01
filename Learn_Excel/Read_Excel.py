# Step 01: install openpyxl: pip install openpyxl

# Read the data from excel
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Book.xlsx')

# return all sheet name of given workbook
print(wb.sheetnames)

# return active sheet name of given workbook
print(wb.active)

# Creating obj frm given sheet name
sh = wb['Dev']

# for row count
print('rows count', sh.max_row)

# for column count
print('column count', sh.max_column)

for row in range(1, sh.max_row+1):
    for column in range(1,sh.max_column+1):
        print(sh.cell(row, column).value)


